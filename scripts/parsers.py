"""纯解析函数：Excel 能力矩阵与日报索引。无 I/O 副作用之外的运行产物，便于测试。"""

from __future__ import annotations

import csv
import hashlib
import re
from collections import Counter
from pathlib import Path

# .NET 风格时间戳（7 位小数）无法直接 fromisoformat，先截断到微秒
_FRAC = re.compile(r"(\.\d{6})\d+")
_FILENAME_DATE = re.compile(r"^(\d{4}-\d{2}-\d{2})-(.+)\.md$")


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def parse_ts(raw: str) -> tuple[str | None, str | None]:
    """归一化时间戳并校验可解析；失败返回 (None, 问题标记)。"""
    from datetime import datetime

    try:
        normalized = _FRAC.sub(r"\1", raw.strip())
        datetime.fromisoformat(normalized)
    except (AttributeError, ValueError):
        return None, "published_at_invalid"
    return normalized, None


def parse_matrix(xlsx: Path) -> dict:
    """解析数字经济岗位能力矩阵：46 岗位 x 30 能力 x 7 分组 + 评分图例。"""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["岗位能力矩阵"]
    rows = list(ws.iter_rows(values_only=True))
    head1, head2 = rows[0], rows[1]

    # 表头两行：第一行能力分组（合并单元格，向右填充），第二行是能力名
    group = None
    capabilities = []
    for col in range(3, len(head2)):
        if col < len(head1) and head1[col]:
            group = str(head1[col]).strip()
        name = head2[col]
        if not name:
            continue
        capabilities.append(
            {
                "capability_id": f"cap_{len(capabilities) + 1:02d}",
                "group": group,
                "name": str(name).strip(),
            }
        )
    cap_ids = [c["capability_id"] for c in capabilities]

    positions, issues = [], []
    for n, row in enumerate(rows[2:], start=1):
        if not row or not row[1]:
            continue
        scores = {}
        for cap_id, value in zip(cap_ids, row[3 : 3 + len(cap_ids)]):
            try:
                scores[cap_id] = int(value) if value is not None else None
            except (TypeError, ValueError):
                scores[cap_id] = None
                issues.append(f"pos_{n:02d}:{cap_id} 无法解析的分值 {value!r}")
        bad = [c for c, v in scores.items() if v is None or not 0 <= v <= 5]
        if bad:
            issues.append(f"pos_{n:02d} 缺失或越界分值 {len(bad)} 项")
        prev_group = positions[-1]["group"] if positions else ""
        positions.append(
            {
                "position_id": f"pos_{n:02d}",
                "group": str(row[0]).strip() if row[0] else prev_group,
                "name": str(row[1]).strip(),
                "summary": str(row[2]).strip() if row[2] else "",
                "scores": scores,
            }
        )

    legend_ws = wb["图例与说明"]
    legend = {
        str(r[0]).strip(): str(r[1]).strip()
        for r in legend_ws.iter_rows(values_only=True)
        if r[0] is not None and str(r[0]).strip().isdigit()
    }
    return {
        "capabilities": capabilities,
        "positions": positions,
        "groups": sorted({p["group"] for p in positions}),
        "score_legend": legend,
        "issues": issues,
    }


def stage_wechat(base: Path) -> dict:
    """以 jueya-index.csv 为权威元数据整理日报归档。

    返回索引内条目（reports）与磁盘上存在但索引未登记的早期文件（unindexed），
    两类都保留 content_hash 与相对路径，正文不复制。
    """
    index_rows = list(csv.DictReader((base / "jueya-index.csv").open(encoding="utf-8-sig")))

    items, seen_ids, issues = [], Counter(), []
    referenced = set()
    for row in index_rows:
        item_id = f"wechat-mp:{row['app_msg_id']}:{row['item_idx']}"
        seen_ids[item_id] += 1
        rel_path = (row.get("markdown_path") or "").replace("\\", "/")
        issues_row = []
        if not rel_path:
            # 索引登记了文章但正文从未保存成功
            issues_row.append("body_not_saved")
            size, digest, path_out = 0, None, ""
        else:
            rel = (base / rel_path).resolve()
            referenced.add(str(rel))
            if not rel.is_file():
                issues_row.append("file_missing")
                size, digest, path_out = 0, None, rel_path
            else:
                size = rel.stat().st_size
                digest = sha256_of(rel)
                path_out = str(rel)
                if size == 0:
                    issues_row.append("empty_file")
        if row["body_status"] != "succeeded":
            issues_row.append(f"body_status={row['body_status']}")
        published, ts_issue = parse_ts(row["published_at"])
        if ts_issue:
            issues_row.append(ts_issue)
        if seen_ids[item_id] > 1:
            issues_row.append("duplicate_item_id")
        items.append(
            {
                "item_id": item_id,
                "source_id": "wechat-mp",
                "guid": f"{row['app_msg_id']}:{row['item_idx']}",
                "title": row["title"],
                "published_at": published,
                "available_at": published,
                "content_hash": digest,
                "ingestion_mode": "backfill",
                "status": "ok" if not issues_row else "quarantined",
                "path": path_out,
                "size_bytes": size,
                "issues": issues_row,
            }
        )
        if issues_row:
            issues.append(f"{item_id}: {','.join(issues_row)}")

    unindexed = []
    for p in sorted(base.rglob("*.md")):
        if str(p.resolve()) in referenced:
            continue
        m = _FILENAME_DATE.match(p.name)
        unindexed.append(
            {
                "item_id": f"wechat-mp:unindexed:{sha256_of(p)[:12]}",
                "source_id": "wechat-mp",
                "guid": None,
                "title": m.group(2) if m else p.stem,
                "published_at": None,
                "published_date": m.group(1) if m else None,
                "available_at": None,
                "content_hash": sha256_of(p),
                "ingestion_mode": "backfill",
                "status": "ok_unindexed",
                "path": str(p),
                "size_bytes": p.stat().st_size,
                "issues": ["not_in_index"],
            }
        )

    def month(item: dict) -> str:
        return (item.get("published_at") or item.get("published_date") or "?")[:7]

    months = Counter(month(i) for i in items + unindexed if month(i) != "?")
    return {
        "reports": items,
        "unindexed": unindexed,
        "issues": issues,
        "metrics": {
            "index_rows": len(items),
            "ok": sum(1 for i in items if i["status"] == "ok"),
            "body_not_saved": sum(1 for i in items if "body_not_saved" in i["issues"]),
            "quarantined": sum(1 for i in items if i["status"] == "quarantined"),
            "unindexed_files": len(unindexed),
            "first_month": min(months) if months else None,
            "last_month": max(months) if months else None,
        },
    }
